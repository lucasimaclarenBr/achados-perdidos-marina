import io
import re
import streamlit as st
import pandas as pd
from datetime import datetime
from infra.banco_dados import supabase
from infra.utils import agora_brt, hoje_brt, formatar_dt_brt

# =============================================================================
# CONSTANTES
# =============================================================================
CATEGORIAS = [
    "Acessórios", "Bolas", "Bolsas", "Brinquedos", "Calçados",
    "Cosméticos", "Garrafas", "Material Esportivo", "Objetos",
    "Óculos", "Roupas", "Toalhas",
]

STATUS_OPCOES = ["Aberto", "Encontrado", "Encerrado"]

COLUNAS_EXPORT = [
    "id_ticket", "nome_socio", "titulo_socio", "telefone", "categoria",
    "descricao_perdido", "data_registro", "data_limite_retorno",
    "status_busca", "dias_no_status",
]

CABECALHOS_EXPORT = [
    "ID", "Nome Sócio", "Título", "Telefone", "Categoria",
    "Descrição", "Data Registro", "Data Limite Retorno",
    "Status", "Dias no Status",
]

PERFIS_EDICAO = {"Admin", "Edicao"}


# =============================================================================
# FUNÇÕES DE DADOS
# =============================================================================

@st.cache_data(ttl=20)
def _carregar_busca_ativa() -> pd.DataFrame:
    resp = supabase.table("busca_ativa").select("*").execute()
    if not resp.data:
        return pd.DataFrame()
    return pd.DataFrame(resp.data)


@st.cache_data(ttl=20)
def _carregar_datas_ultimo_status() -> dict:
    try:
        resp = (
            supabase.table("historico_busca_ativa")
            .select("id_ticket, data_hora_alteracao")
            .eq("campo_alterado", "status_busca")
            .order("data_hora_alteracao", desc=True)
            .execute()
        )
        visto = {}
        for row in resp.data:
            tid = row["id_ticket"]
            if tid not in visto:
                visto[tid] = row["data_hora_alteracao"]
        return visto
    except Exception:
        return {}


def _enriquecer_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    datas_status = _carregar_datas_ultimo_status()
    hoje = hoje_brt()

    def _dias(row):
        tid = int(row["id_ticket"])
        if tid in datas_status:
            dt = datas_status[tid]
            if isinstance(dt, str):
                dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
            return (hoje - dt.date()).days
        if pd.notna(row.get("data_registro")):
            dc = row["data_registro"]
            if isinstance(dc, str):
                dc = datetime.fromisoformat(dc.replace("Z", "+00:00"))
            return (hoje - dc.date()).days
        return 0

    df["dias_no_status"] = df.apply(_dias, axis=1)
    df["data_registro"] = pd.to_datetime(df["data_registro"], utc=True, errors="coerce")
    df["data_registro_fmt"] = df["data_registro"].dt.strftime("%d/%m/%Y").fillna("")
    df["data_limite_retorno"] = pd.to_datetime(df["data_limite_retorno"], errors="coerce")
    df["data_limite_retorno_fmt"] = df["data_limite_retorno"].dt.strftime("%d/%m/%Y").fillna("")
    return df


def _aplicar_filtros(df: pd.DataFrame, filtros: dict) -> pd.DataFrame:
    if filtros.get("categorias"):
        df = df[df["categoria"].isin(filtros["categorias"])]
    if filtros.get("status"):
        df = df[df["status_busca"].isin(filtros["status"])]

    datas = filtros.get("datas", ())
    if isinstance(datas, (tuple, list)) and len(datas) == 2:
        df = df[
            (df["data_registro"].dt.date >= datas[0]) &
            (df["data_registro"].dt.date <= datas[1])
        ]
    elif filtros.get("dias"):
        df = df[df["dias_no_status"] <= filtros["dias"]]

    data_limite = filtros.get("data_limite", ())
    if isinstance(data_limite, (tuple, list)) and len(data_limite) == 2:
        df = df[
            (df["data_limite_retorno"].dt.date >= data_limite[0]) &
            (df["data_limite_retorno"].dt.date <= data_limite[1])
        ]

    if filtros.get("texto"):
        from unidecode import unidecode
        t = unidecode(filtros["texto"].lower())
        df = df[
            df["nome_socio"].fillna("").apply(lambda x: unidecode(x.lower())).str.contains(t, na=False) |
            df["descricao_perdido"].fillna("").apply(lambda x: unidecode(x.lower())).str.contains(t, na=False)
        ]

    if filtros.get("titulo"):
        t_norm = re.sub(r"\D", "", filtros["titulo"])
        if t_norm:
            df = df[
                df["titulo_socio"].fillna("").apply(lambda x: re.sub(r"\D", "", x)).str.contains(t_norm, na=False)
            ]
    return df


# =============================================================================
# EXPORTAÇÃO
# =============================================================================

def _preparar_df_export(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in COLUNAS_EXPORT if c in df.columns]
    df_exp = df[cols].copy()
    cabecalhos = CABECALHOS_EXPORT[: len(cols)]
    df_exp.columns = cabecalhos
    return df_exp


def _exportar_excel(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl import Workbook

    df_exp = _preparar_df_export(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Busca Ativa"

    header_fill = PatternFill("solid", start_color="002366")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)

    for col_idx, col_name in enumerate(df_exp.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df_exp.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="F0F4FF")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sanitizar_pdf(texto: str) -> str:
    """Remove/substitui caracteres fora do Latin-1 para compatibilidade com Helvetica."""
    substituicoes = {
        "—": "-", "–": "-", "‘": "'", "’": "'",
        "“": '"', "”": '"', "…": "...", "ç": "c",
        "ã": "a", "õ": "o", "á": "a", "é": "e",
        "í": "i", "ó": "o", "ú": "u", "à": "a",
        "â": "a", "ê": "e", "ô": "o", "ü": "u",
        "ñ": "n", "Ç": "C", "Ã": "A", "Õ": "O",
        "Á": "A", "É": "E", "Í": "I", "Ó": "O",
        "Ú": "U", "À": "A", "Â": "A", "Ê": "E",
        "Ô": "O",
    }
    for char, sub in substituicoes.items():
        texto = texto.replace(char, sub)
    return texto.encode("latin-1", errors="ignore").decode("latin-1")


def _exportar_pdf(df: pd.DataFrame) -> bytes | None:
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    df_exp = _preparar_df_export(df)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_margins(10, 10, 10)

    pdf.set_fill_color(0, 35, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)

    n_cols = len(df_exp.columns)
    col_w = 267 / n_cols

    for h in df_exp.columns:
        pdf.cell(col_w, 8, _sanitizar_pdf(str(h)[:20]), border=0, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(30, 30, 30)
    pdf.set_font("Helvetica", size=7)

    for i, (_, row) in enumerate(df_exp.iterrows()):
        if i % 2 == 0:
            pdf.set_fill_color(240, 244, 255)
        else:
            pdf.set_fill_color(255, 255, 255)
        for col in df_exp.columns:
            val = _sanitizar_pdf(str(row[col] if pd.notna(row[col]) else "")[:24])
            pdf.cell(col_w, 6, val, border=0, fill=True)
        pdf.ln()

    pdf.set_y(-15)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(150, 150, 150)
    rodape = _sanitizar_pdf(f"Marina Barra Clube - Achados e Perdidos - Gerado em {hoje_brt().strftime('%d/%m/%Y')}")
    pdf.cell(0, 10, rodape, align="C")

    return bytes(pdf.output())


# =============================================================================
# DIALOG DE DETALHE / EDIÇÃO
# =============================================================================

@st.dialog("Detalhes e Edição do Ticket", width="large")
def _dialog_ticket(ticket: dict, perfil: str, usuario: str):
    from telas.tela_cadastro import _formatar_telefone, _formatar_titulo, _validar_telefone, _validar_titulo

    pode_editar = perfil in PERFIS_EDICAO
    tid = ticket["id_ticket"]
    cod = str(tid)

    st.markdown(f"**Ticket:** `#{tid}`")
    st.markdown(f"**Registrado em:** {ticket.get('data_registro_fmt', '')}")
    st.markdown(f"**Data limite de retorno:** {ticket.get('data_limite_retorno_fmt', '')}")
    st.markdown(f"**Dias no status:** {ticket.get('dias_no_status', 0)}")

    st.markdown("---")

    if not pode_editar:
        st.markdown(f"**Sócio:** {ticket.get('nome_socio', '') or '—'}")
        st.markdown(f"**Título:** {ticket.get('titulo_socio', '') or '—'}")
        st.markdown(f"**Telefone:** {ticket.get('telefone', '') or '—'}")
        st.markdown(f"**Categoria:** {ticket.get('categoria', '') or '—'}")
        st.markdown(f"**Descrição:** {ticket.get('descricao_perdido', '') or '—'}")
        st.markdown(f"**Status:** {ticket.get('status_busca', '') or '—'}")
    else:
        def _upper_nome():
            st.session_state[f"ed_nome_{cod}"] = st.session_state[f"ed_nome_{cod}"].upper()

        novo_nome = st.text_input(
            "Nome do Sócio",
            value=ticket.get("nome_socio", ""),
            key=f"ed_nome_{cod}",
            on_change=_upper_nome,
        )

        sc1, sc2 = st.columns(2)

        with sc1:
            def _fmt_titulo():
                st.session_state[f"ed_titulo_{cod}"] = _formatar_titulo(st.session_state[f"ed_titulo_{cod}"])

            novo_titulo = st.text_input(
                "Título",
                value=ticket.get("titulo_socio", ""),
                key=f"ed_titulo_{cod}",
                on_change=_fmt_titulo,
                placeholder="1234-10",
            )
            if novo_titulo and not _validar_titulo(novo_titulo):
                digitos = len(re.sub(r"\D", "", novo_titulo))
                if digitos > 6:
                    st.caption(f"⚠️ Título muito longo ({digitos} dígitos). Use o padrão 1234-10.")
                else:
                    st.caption(f"⚠️ Título incompleto ({digitos} dígitos). Use o padrão 1234-10.")

        with sc2:
            def _fmt_tel():
                st.session_state[f"ed_tel_{cod}"] = _formatar_telefone(st.session_state[f"ed_tel_{cod}"])

            novo_tel = st.text_input(
                "Telefone",
                value=ticket.get("telefone", ""),
                key=f"ed_tel_{cod}",
                on_change=_fmt_tel,
                placeholder="(21) 99999-9999",
            )
            if novo_tel:
                digitos = len(re.sub(r"\D", "", novo_tel))
                if digitos > 0 and digitos != 11:
                    st.caption(f"⚠️ Número inválido ({digitos} dígitos). Use DDD + 9 dígitos.")

        sc3, sc4 = st.columns(2)

        with sc3:
            categorias_idx = CATEGORIAS.index(ticket["categoria"]) if ticket.get("categoria") in CATEGORIAS else 0
            nova_categoria = st.selectbox(
                "Categoria",
                options=CATEGORIAS,
                index=categorias_idx,
                key=f"ed_cat_{cod}",
            )

        with sc4:
            status_idx = STATUS_OPCOES.index(ticket["status_busca"]) if ticket.get("status_busca") in STATUS_OPCOES else 0
            novo_status = st.selectbox(
                "Status",
                options=STATUS_OPCOES,
                index=status_idx,
                key=f"ed_status_{cod}",
            )

        def _upper_desc():
            st.session_state[f"ed_desc_{cod}"] = st.session_state[f"ed_desc_{cod}"].upper()

        nova_descricao = st.text_input(
            "Descrição do que foi perdido",
            value=ticket.get("descricao_perdido", ""),
            key=f"ed_desc_{cod}",
            on_change=_upper_desc,
        )

    # ── Histórico de edições ──
    st.markdown("---")
    st.markdown('<p class="section-header">Histórico de Edições</p>', unsafe_allow_html=True)
    try:
        resp_hist = (
            supabase.table("historico_busca_ativa")
            .select("campo_alterado, valor_antigo, valor_novo, data_hora_alteracao, usuario_editor")
            .eq("id_ticket", tid)
            .order("data_hora_alteracao", desc=True)
            .execute()
        )
        if resp_hist.data:
            for h in resp_hist.data:
                dt = formatar_dt_brt(h["data_hora_alteracao"])
                st.caption(f"🔄 {h['valor_antigo']} → **{h['valor_novo']}** — {dt} por *{h['usuario_editor']}*")
        else:
            st.caption("Nenhuma movimentação registrada.")
    except Exception as e:
        st.caption(f"Erro ao carregar histórico: {e}")

    if not pode_editar:
        return

    st.markdown("---")
    senha_conf = st.text_input(
        "Senha para confirmar alterações",
        type="password",
        key=f"ed_senha_{cod}",
    )

    col_s, col_c = st.columns(2)
    with col_s:
        salvar_clicado = st.button("Salvar alterações", use_container_width=True, key=f"ed_salvar_{cod}")
    with col_c:
        if st.button("Descartar", use_container_width=True, key=f"ed_cancelar_{cod}"):
            st.rerun()

    if salvar_clicado:
        if not senha_conf:
            st.toast("Digite sua senha para confirmar.", icon="⚠️")
        else:
            try:
                resp = supabase.table("usuarios").select("senha").eq("login", usuario).execute()
                if not resp.data or resp.data[0]["senha"] != senha_conf:
                    st.toast("Senha incorreta.", icon="❌")
                else:
                    atualizacoes = {}
                    historico = []
                    agora = agora_brt().isoformat()

                    campos = [
                        ("nome_socio", ticket.get("nome_socio", ""), novo_nome),
                        ("titulo_socio", ticket.get("titulo_socio", ""), novo_titulo),
                        ("telefone", ticket.get("telefone", ""), novo_tel),
                        ("categoria", ticket.get("categoria", ""), nova_categoria),
                        ("descricao_perdido", ticket.get("descricao_perdido", ""), nova_descricao),
                        ("status_busca", ticket.get("status_busca", ""), novo_status),
                    ]
                    for campo, antigo, novo in campos:
                        if novo != antigo:
                            atualizacoes[campo] = novo
                            historico.append({
                                "id_ticket": tid, "campo_alterado": campo,
                                "valor_antigo": str(antigo), "valor_novo": str(novo),
                                "data_hora_alteracao": agora, "usuario_editor": usuario,
                            })

                    if not atualizacoes:
                        st.toast("Nenhuma alteração detectada.", icon="ℹ️")
                    else:
                        try:
                            supabase.table("busca_ativa").update(atualizacoes).eq("id_ticket", tid).execute()
                            if historico:
                                supabase.table("historico_busca_ativa").insert(historico).execute()
                            _carregar_busca_ativa.clear()
                            _carregar_datas_ultimo_status.clear()
                            st.toast(f"Ticket #{tid} atualizado!", icon="✅")
                            st.rerun()
                        except Exception as e:
                            st.toast(f"Erro ao salvar: {e}", icon="❌")
            except Exception as e:
                st.toast(f"Erro ao validar senha: {e}", icon="❌")


# =============================================================================
# TELA PRINCIPAL
# =============================================================================

def mostrar_tela():
    st.title("Busca Ativa — Achados e Perdidos")

    perfil = st.session_state.get("perfil", "Consulta")
    usuario = st.session_state.get("login", "")

    df_raw = _carregar_busca_ativa()
    if df_raw.empty:
        st.info("Nenhum ticket de busca ativa registrado ainda.")
        return
    df = _enriquecer_df(df_raw.copy())

    # ── Filtros ──
    with st.expander("🔍  Filtros", expanded=True):

        def _ao_marcar_7d():
            st.session_state["fba_14d"] = False
            st.session_state["fba_28d"] = False
            st.session_state["fba_datas"] = ()

        def _ao_marcar_14d():
            st.session_state["fba_7d"] = False
            st.session_state["fba_28d"] = False
            st.session_state["fba_datas"] = ()

        def _ao_marcar_28d():
            st.session_state["fba_7d"] = False
            st.session_state["fba_14d"] = False
            st.session_state["fba_datas"] = ()

        def _ao_alterar_datas():
            val = st.session_state.get("fba_datas", ())
            if isinstance(val, (tuple, list)) and len(val) == 2:
                st.session_state["fba_7d"] = False
                st.session_state["fba_14d"] = False
                st.session_state["fba_28d"] = False

        col_esq, col_ctr, col_dir = st.columns([4, 1, 3])

        with col_esq:
            f_texto = st.text_input(
                "Buscar (nome do sócio, descrição)",
                placeholder="Digite aqui...",
                key="fba_texto",
            )
            f_titulo = st.text_input(
                "Código do Título",
                placeholder="1234-10",
                key="fba_titulo",
            )
            st.markdown('<label style="font-size:0.875rem; font-weight:400; color:#fafafa">Contagem de Dias (Opcional)</label>', unsafe_allow_html=True)
            cd1, cd2, cd3, cd4 = st.columns([1, 1, 1, 2])
            with cd1:
                f_7d = st.checkbox("7 Dias", key="fba_7d", on_change=_ao_marcar_7d)
            with cd2:
                f_14d = st.checkbox("14 Dias", key="fba_14d", on_change=_ao_marcar_14d)
            with cd3:
                f_28d = st.checkbox("28 Dias", key="fba_28d", on_change=_ao_marcar_28d)
            with cd4:
                f_datas = st.date_input(
                    "Personalizado",
                    value=(),
                    key="fba_datas",
                    on_change=_ao_alterar_datas,
                    label_visibility="collapsed",
                )

        with col_dir:
            f_cats = st.multiselect(
                "Categoria (Opcional)",
                options=CATEGORIAS,
                placeholder="Selecione a(s) categoria(s)",
                key="fba_cats",
            )
            f_status = st.multiselect(
                "Status (Opcional)",
                options=STATUS_OPCOES,
                placeholder="Selecione o(s) status",
                key="fba_status",
            )
            f_data_limite = st.date_input(
                "Data limite de retorno",
                value=(),
                key="fba_data_limite",
            )

    dias_filtro = 7 if f_7d else (14 if f_14d else (28 if f_28d else None))

    df_filtrado = _aplicar_filtros(df.copy(), {
        "categorias": f_cats,
        "status": f_status,
        "texto": f_texto,
        "titulo": f_titulo,
        "dias": dias_filtro,
        "datas": f_datas,
        "data_limite": f_data_limite,
    })

    st.caption(f"{len(df_filtrado)} ticket(s) encontrado(s)")

    if df_filtrado.empty:
        st.info("Nenhum ticket corresponde aos filtros.")
        return

    # ── Tabela ──
    df_editor = df_filtrado[
        ["id_ticket", "nome_socio", "titulo_socio", "descricao_perdido", "categoria", "status_busca", "dias_no_status"]
    ].copy()
    df_editor.columns = ["ID", "Nome Sócio", "Título", "Descrição", "Categoria", "Status", "Dias"]
    df_editor.insert(0, "✓", False)

    resultado = st.data_editor(
        df_editor,
        use_container_width=True,
        hide_index=True,
        column_config={
            "✓":          st.column_config.CheckboxColumn("✓", width="small"),
            "ID":         st.column_config.NumberColumn("ID", width="small"),
            "Nome Sócio": st.column_config.TextColumn("Nome Sócio", width="medium"),
            "Título":     st.column_config.TextColumn("Título", width="small"),
            "Descrição":  st.column_config.TextColumn("Descrição", width="large"),
            "Categoria":  st.column_config.TextColumn("Categoria", width="medium"),
            "Status":     st.column_config.TextColumn("Status", width="medium"),
            "Dias":       st.column_config.NumberColumn("Dias", width="small"),
        },
        disabled=["ID", "Nome Sócio", "Título", "Descrição", "Categoria", "Status", "Dias"],
        key="tabela_busca_ativa",
    )

    indices_marcados = resultado.reset_index(drop=True)
    indices_marcados = indices_marcados[indices_marcados["✓"] == True].index.tolist()
    selecionados = df_filtrado["id_ticket"].reset_index(drop=True).iloc[indices_marcados].tolist()

    # ── Edição e exportação ──
    st.markdown("---")
    col_editar, col_exp_excel, col_exp_pdf = st.columns([1, 1, 1])

    with col_editar:
        if perfil in PERFIS_EDICAO:
            if st.button(
                "Editar",
                disabled=len(selecionados) != 1,
                use_container_width=True,
                help="Selecione um ticket",
                key="btn_editar_busca",
            ):
                row_sel = df_filtrado[df_filtrado["id_ticket"] == selecionados[0]].iloc[0]
                _dialog_ticket(row_sel.to_dict(), perfil, usuario)

    df_sel = (
        df_filtrado[df_filtrado["id_ticket"].isin(selecionados)]
        if selecionados
        else pd.DataFrame(columns=COLUNAS_EXPORT)
    )

    with col_exp_excel:
        with st.popover("⬇ Exportar para Excel", use_container_width=True):
            st.download_button(
                "Exportar tudo",
                data=_exportar_excel(df_filtrado),
                file_name=f"busca_ativa_{hoje_brt().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_exp_excel_tudo_busca",
            )
            label_sel = f"Exportar seleção ({len(selecionados)})" if selecionados else "Exportar seleção"
            st.download_button(
                label_sel,
                data=_exportar_excel(df_sel),
                file_name=f"busca_ativa_selecao_{hoje_brt().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_exp_excel_sel_busca",
            )

    with col_exp_pdf:
        with st.popover("⬇ Exportar para PDF", use_container_width=True):
            pdf_tudo = _exportar_pdf(df_filtrado)
            if pdf_tudo:
                st.download_button(
                    "Exportar tudo",
                    data=pdf_tudo,
                    file_name=f"busca_ativa_{hoje_brt().isoformat()}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_exp_pdf_tudo_busca",
                )
            else:
                st.caption("Instale fpdf2 para habilitar PDF.")
            pdf_sel = _exportar_pdf(df_sel)
            if pdf_sel:
                label_sel_pdf = f"Exportar seleção ({len(selecionados)})" if selecionados else "Exportar seleção"
                st.download_button(
                    label_sel_pdf,
                    data=pdf_sel,
                    file_name=f"busca_ativa_selecao_{hoje_brt().isoformat()}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_exp_pdf_sel_busca",
                )
            else:
                st.caption("Instale fpdf2 para habilitar PDF.")
