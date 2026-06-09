import io
import re
import streamlit as st
import pandas as pd
from datetime import date, datetime
from infra.banco_dados import supabase


# =============================================================================
# CONSTANTES
# =============================================================================
STATUS_OPCOES = ["Armazenado", "Caixa Azul", "Museu", "A Doar", "Doados", "Descartados", "Devolvidos"]

TRANSICOES = {
    "Armazenado":   ["Museu", "Devolvidos", "A Doar", "Descartados"],
    "Museu":       ["A Doar", "Descartados", "Devolvidos"],
    "A Doar":      ["Doados"],
    "Doados":      [],
    "Descartados": [],
    "Devolvidos":  [],
    "Caixa Azul":  [],
}

CATEGORIAS = [
    "Acessórios", "Bolas", "Bolsas", "Brinquedos", "Calçados",
    "Cosméticos", "Garrafas", "Material Esportivo", "Objetos",
    "Óculos", "Roupas", "Toalhas",
]

COLUNAS_EXPORT = [
    "codigo_item", "categoria", "descricao", "local_achado", "status_exibicao",
    "dias_no_status", "data_cadastro_fmt", "nome_socio_identificado",
    "titulo_socio", "telefone_socio", "contatado", "caixa_azul",
]

CABECALHOS_EXPORT = [
    "Código", "Categoria", "Descrição", "Local", "Status",
    "Dias no Status", "Data Cadastro", "Nome Sócio",
    "Título", "Telefone", "Contatado", "Caixa Azul",
]

PERFIS_EDICAO = {"Admin", "Edicao"}

# Badges de cor por status (para exibição futura)
COR_STATUS = {
    "Armazenado":   "#16a34a",
    "Caixa Azul":  "#2563eb",
    "Museu":       "#d97706",
    "A Doar":      "#7c3aed",
    "Doados":      "#64748b",
    "Descartados": "#dc2626",
    "Devolvidos":  "#64748b",
}


# =============================================================================
# FUNÇÕES DE DADOS
# =============================================================================

@st.cache_data(ttl=20)
def _carregar_itens() -> pd.DataFrame:
    resp = supabase.table("itens").select("*").execute()
    if not resp.data:
        return pd.DataFrame()
    return pd.DataFrame(resp.data)


@st.cache_data(ttl=20)
def _carregar_datas_ultimo_status() -> dict:
    try:
        resp = (
            supabase.table("historico_edicoes")
            .select("codigo_item, data_hora_alteracao")
            .eq("campo_alterado", "status_atual")
            .order("data_hora_alteracao", desc=True)
            .execute()
        )
        visto = {}
        for row in resp.data:
            cod = row["codigo_item"]
            if cod not in visto:
                visto[cod] = row["data_hora_alteracao"]
        return visto
    except Exception:
        return {}


def _enriquecer_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    datas_status = _carregar_datas_ultimo_status()
    hoje = date.today()

    def _dias(row):
        cod = row["codigo_item"]
        if cod in datas_status:
            dt = datas_status[cod]
            if isinstance(dt, str):
                dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
            return (hoje - dt.date()).days
        if pd.notna(row.get("data_cadastro")):
            dc = row["data_cadastro"]
            if isinstance(dc, str):
                dc = datetime.fromisoformat(dc.replace("Z", "+00:00"))
            return (hoje - dc.date()).days
        return 0

    df["dias_no_status"] = df.apply(_dias, axis=1)

    def _status_exib(row):
        if row.get("caixa_azul") and row.get("status_atual") == "Armazenado":
            return "Caixa Azul"
        return row.get("status_atual", "")

    df["status_exibicao"] = df.apply(_status_exib, axis=1)
    df["data_cadastro"] = pd.to_datetime(df["data_cadastro"], utc=True, errors="coerce")
    df["data_cadastro_fmt"] = df["data_cadastro"].dt.strftime("%d/%m/%Y").fillna("")
    return df


def _aplicar_filtros(df: pd.DataFrame, filtros: dict) -> pd.DataFrame:
    if filtros.get("categoria"):
        df = df[df["categoria"] == filtros["categoria"]]
    if filtros.get("status"):
        df = df[df["status_exibicao"] == filtros["status"]]
    if filtros.get("caixa_azul"):
        df = df[df["caixa_azul"] == True]

    # Datas personalizado tem prioridade sobre filtro de dias
    datas = filtros.get("datas", ())
    if isinstance(datas, (tuple, list)) and len(datas) == 2:
        df = df[
            (df["data_cadastro"].dt.date >= datas[0]) &
            (df["data_cadastro"].dt.date <= datas[1])
        ]
    elif filtros.get("dias"):
        df = df[df["dias_no_status"] <= filtros["dias"]]

    if filtros.get("texto"):
        t = filtros["texto"].lower()
        df = df[
            df["descricao"].str.lower().str.contains(t, na=False) |
            df["codigo_item"].str.lower().str.contains(t, na=False) |
            df["nome_socio_identificado"].fillna("").str.lower().str.contains(t, na=False)
        ]
    return df


# =============================================================================
# EXPORTAÇÃO
# =============================================================================

def _preparar_df_export(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in COLUNAS_EXPORT if c in df.columns]
    df_exp = df[cols].copy()
    cabecalhos = CABECALHOS_EXPORT[: len(cols)]
    df_exp.columns = cabecalhos
    return df_exp


def _exportar_excel(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl import Workbook

    df_exp = _preparar_df_export(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventário"

    header_fill = PatternFill("solid", start_color="002366")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)

    for col_idx, col_name in enumerate(df_exp.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df_exp.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="F0F4FF")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sanitizar_pdf(texto: str) -> str:
    """Remove/substitui caracteres fora do Latin-1 para compatibilidade com Helvetica."""
    substituicoes = {
        "—": "-", "–": "-", "‘": "'", "’": "'",
        "“": '"', "”": '"', "…": "...", "ç": "c",
        "ã": "a", "õ": "o", "á": "a", "é": "e",
        "í": "i", "ó": "o", "ú": "u", "à": "a",
        "â": "a", "ê": "e", "ô": "o", "ü": "u",
        "ñ": "n", "Ç": "C", "Ã": "A", "Õ": "O",
        "Á": "A", "É": "E", "Í": "I", "Ó": "O",
        "Ú": "U", "À": "A", "Â": "A", "Ê": "E",
        "Ô": "O",
    }
    for char, sub in substituicoes.items():
        texto = texto.replace(char, sub)
    # Remove qualquer outro caractere fora do Latin-1
    return texto.encode("latin-1", errors="ignore").decode("latin-1")


def _exportar_pdf(df: pd.DataFrame) -> bytes | None:
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    df_exp = _preparar_df_export(df)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_margins(10, 10, 10)

    # Cabeçalho
    pdf.set_fill_color(0, 35, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)

    n_cols = len(df_exp.columns)
    col_w = 267 / n_cols

    for h in df_exp.columns:
        pdf.cell(col_w, 8, _sanitizar_pdf(str(h)[:20]), border=0, fill=True, align="C")
    pdf.ln()

    # Linhas
    pdf.set_text_color(30, 30, 30)
    pdf.set_font("Helvetica", size=7)

    for i, (_, row) in enumerate(df_exp.iterrows()):
        if i % 2 == 0:
            pdf.set_fill_color(240, 244, 255)
        else:
            pdf.set_fill_color(255, 255, 255)
        for col in df_exp.columns:
            val = _sanitizar_pdf(str(row[col] if pd.notna(row[col]) else "")[:24])
            pdf.cell(col_w, 6, val, border=0, fill=True)
        pdf.ln()

    # Rodapé
    pdf.set_y(-15)
    pdf.set_font("Helvetica", "I", 7)
    pdf.set_text_color(150, 150, 150)
    rodape = _sanitizar_pdf(f"Marina Barra Clube - Achados e Perdidos - Gerado em {date.today().strftime('%d/%m/%Y')}")
    pdf.cell(0, 10, rodape, align="C")

    return bytes(pdf.output())


# =============================================================================
# DIALOG DE DETALHE / EDIÇÃO
# =============================================================================

@st.dialog("Detalhes e Edição do Item", width="large")
def _dialog_item(item: dict, perfil: str, usuario: str):
    pode_editar = perfil in PERFIS_EDICAO
    cod = item["codigo_item"]
    status_atual = item.get("status_atual", "Armazenado")

    col_foto, col_dados = st.columns([1, 2])

    with col_foto:
        if item.get("foto_url"):
            st.image(item["foto_url"], use_container_width=True)
        else:
            st.markdown(
                "<div style='background:#f1f5f9;border-radius:8px;height:200px;"
                "display:flex;align-items:center;justify-content:center;"
                "color:#94a3b8;font-size:0.85rem'>Sem foto</div>",
                unsafe_allow_html=True,
            )
        st.markdown(f"**Código:** `{cod}`")
        st.markdown(f"**Categoria:** {item.get('categoria', '')}")
        st.markdown(f"**Cadastrado:** {item.get('data_cadastro_fmt', '')}")
        st.markdown(f"**Dias no status:** {item.get('dias_no_status', 0)}")

    with col_dados:
        if not pode_editar:
            st.markdown(f"**Descrição:** {item.get('descricao', '')}")
            st.markdown(f"**Local:** {item.get('local_achado', '') or '—'}")
            st.markdown(f"**Status:** {item.get('status_exibicao', '')}")
            st.markdown(f"**Caixa Azul:** {'Sim' if item.get('caixa_azul') else 'Não'}")
            st.markdown("---")
            st.markdown(f"**Sócio:** {item.get('nome_socio_identificado', '') or '—'}")
            st.markdown(f"**Título:** {item.get('titulo_socio', '') or '—'}")
            st.markdown(f"**Telefone:** {item.get('telefone_socio', '') or '—'}")
            st.markdown(f"**Contatado:** {'Sim' if item.get('contatado') else 'Não'}")
            return

        nova_descricao = st.text_input("Descrição", value=item.get("descricao", ""), key=f"ed_desc_{cod}")
        novo_local = st.text_input("Local onde foi achado", value=item.get("local_achado") or "", key=f"ed_local_{cod}")

        opcoes_transicao = TRANSICOES.get(status_atual, [])
        if opcoes_transicao:
            novo_status = st.selectbox(
                f"Status atual: **{item.get('status_exibicao', status_atual)}** → Mover para",
                options=["Manter atual"] + opcoes_transicao,
                key=f"ed_status_{cod}",
            )
        else:
            st.markdown(f"**Status:** {item.get('status_exibicao', '')} *(status final)*")
            novo_status = "Manter atual"

        nova_caixa_azul = item.get("caixa_azul", False)
        if status_atual == "Armazenado":
            nova_caixa_azul = st.checkbox("Caixa Azul", value=nova_caixa_azul, key=f"ed_caixa_{cod}")

        st.markdown("---")
        st.markdown('<p class="section-header">Dados do Sócio</p>', unsafe_allow_html=True)
        sc1, sc2 = st.columns(2)
        with sc1:
            novo_nome_socio = st.text_input("Nome", value=item.get("nome_socio_identificado") or "", key=f"ed_nome_s_{cod}")
            novo_titulo = st.text_input("Título", value=item.get("titulo_socio") or "", key=f"ed_titulo_{cod}")
        with sc2:
            novo_tel = st.text_input("Telefone", value=item.get("telefone_socio") or "", key=f"ed_tel_{cod}")
            novo_contatado = st.checkbox("Sócio contatado", value=bool(item.get("contatado")), key=f"ed_cont_{cod}")

    st.markdown("---")
    senha_conf = st.text_input("Senha para confirmar alterações", type="password", key=f"ed_senha_{cod}")

    col_s, col_c = st.columns(2)
    with col_s:
        if st.button("Salvar alterações", use_container_width=True, key=f"ed_salvar_{cod}"):
            if not senha_conf:
                st.toast("Digite sua senha para confirmar.", icon="⚠️")
                return
            try:
                resp = supabase.table("usuarios").select("senha").eq("login", usuario).execute()
                if not resp.data or resp.data[0]["senha"] != senha_conf:
                    st.toast("Senha incorreta.", icon="❌")
                    return
            except Exception as e:
                st.toast(f"Erro ao validar senha: {e}", icon="❌")
                return

            atualizacoes = {}
            historico = []
            agora = datetime.utcnow().isoformat()

            campos = [
                ("descricao", item.get("descricao", ""), nova_descricao),
                ("local_achado", item.get("local_achado") or "", novo_local),
                ("nome_socio_identificado", item.get("nome_socio_identificado") or "", novo_nome_socio),
                ("titulo_socio", item.get("titulo_socio") or "", novo_titulo),
                ("telefone_socio", item.get("telefone_socio") or "", novo_tel),
                ("contatado", bool(item.get("contatado")), novo_contatado),
                ("caixa_azul", item.get("caixa_azul", False), nova_caixa_azul),
            ]
            for campo, antigo, novo in campos:
                if novo != antigo:
                    atualizacoes[campo] = novo
                    historico.append({
                        "codigo_item": cod, "campo_alterado": campo,
                        "valor_antigo": str(antigo), "valor_novo": str(novo),
                        "data_hora_alteracao": agora, "usuario_editor": usuario,
                    })

            if novo_status != "Manter atual":
                atualizacoes["status_atual"] = novo_status
                historico.append({
                    "codigo_item": cod, "campo_alterado": "status_atual",
                    "valor_antigo": status_atual, "valor_novo": novo_status,
                    "data_hora_alteracao": agora, "usuario_editor": usuario,
                })

            if not atualizacoes:
                st.toast("Nenhuma alteração detectada.", icon="ℹ️")
                return

            try:
                supabase.table("itens").update(atualizacoes).eq("codigo_item", cod).execute()
                if historico:
                    supabase.table("historico_edicoes").insert(historico).execute()
                _carregar_itens.clear()
                _carregar_datas_ultimo_status.clear()
                st.toast(f"Item {cod} atualizado!", icon="✅")
                st.rerun()
            except Exception as e:
                st.toast(f"Erro ao salvar: {e}", icon="❌")

    with col_c:
        if st.button("Descartar", use_container_width=True, key=f"ed_cancelar_{cod}"):
            st.rerun()


# =============================================================================
# DIALOG MOVIMENTAÇÃO EM LOTE
# =============================================================================

@st.dialog("Movimentar itens selecionados", width="small")
def _dialog_lote(codigos: list, perfil: str, usuario: str):
    st.markdown(f"**{len(codigos)} item(ns) selecionado(s)**")
    st.caption(", ".join(codigos[:10]) + ("..." if len(codigos) > 10 else ""))

    novo_status = st.selectbox(
        "Mover todos para",
        options=["Selecione..."] + STATUS_OPCOES,
        key="lote_status",
    )
    senha_lote = st.text_input("Senha de confirmação", type="password", key="lote_senha")

    if st.button("Confirmar", use_container_width=True, key="lote_confirmar"):
        if novo_status == "Selecione...":
            st.toast("Selecione um status.", icon="⚠️")
            return
        if not senha_lote:
            st.toast("Digite sua senha.", icon="⚠️")
            return
        try:
            resp = supabase.table("usuarios").select("senha").eq("login", usuario).execute()
            if not resp.data or resp.data[0]["senha"] != senha_lote:
                st.toast("Senha incorreta.", icon="❌")
                return
        except Exception as e:
            st.toast(f"Erro: {e}", icon="❌")
            return

        agora = datetime.utcnow().isoformat()
        df_atual = _carregar_itens()
        historico = []
        try:
            for cod in codigos:
                row = df_atual[df_atual["codigo_item"] == cod]
                status_ant = row.iloc[0]["status_atual"] if not row.empty else ""
                supabase.table("itens").update({"status_atual": novo_status}).eq("codigo_item", cod).execute()
                historico.append({
                    "codigo_item": cod, "campo_alterado": "status_atual",
                    "valor_antigo": status_ant, "valor_novo": novo_status,
                    "data_hora_alteracao": agora, "usuario_editor": usuario,
                })
            if historico:
                supabase.table("historico_edicoes").insert(historico).execute()
            _carregar_itens.clear()
            _carregar_datas_ultimo_status.clear()
            st.toast(f"{len(codigos)} item(ns) movido(s) para '{novo_status}'.", icon="✅")
            st.rerun()
        except Exception as e:
            st.toast(f"Erro: {e}", icon="❌")


# =============================================================================
# TELA PRINCIPAL
# =============================================================================

def mostrar_tela():
    st.title("Pesquisa de Item — Achados e Perdidos")

    perfil = st.session_state.get("perfil", "Consulta")
    usuario = st.session_state.get("login", "")

    df_raw = _carregar_itens()
    if df_raw.empty:
        st.info("Nenhum item cadastrado ainda.")
        return
    df = _enriquecer_df(df_raw.copy())

    # ── Filtros ──
    with st.expander("🔍  Filtros", expanded=True):

        def _ao_marcar_7d():
            st.session_state["f_14d"] = False
            st.session_state["f_28d"] = False
            st.session_state["f_datas"] = ()

        def _ao_marcar_14d():
            st.session_state["f_7d"] = False
            st.session_state["f_28d"] = False
            st.session_state["f_datas"] = ()

        def _ao_marcar_28d():
            st.session_state["f_7d"] = False
            st.session_state["f_14d"] = False
            st.session_state["f_datas"] = ()

        def _ao_alterar_datas():
            val = st.session_state.get("f_datas", ())
            if isinstance(val, (tuple, list)) and len(val) == 2:
                st.session_state["f_7d"] = False
                st.session_state["f_14d"] = False
                st.session_state["f_28d"] = False

        col_esq, col_ctr, col_dir = st.columns([4, 1, 3])

        with col_esq:
            f_texto = st.text_input(
                "Buscar (código, descrição, sócio)",
                placeholder="Digite aqui...",
                key="f_texto",
            )
            st.markdown('<label style="font-size:0.875rem; font-weight:400; color:#fafafa">Contagem de Dias (Opcional)</label>', unsafe_allow_html=True)
            cd1, cd2, cd3, cd4 = st.columns([1, 1, 1, 2])
            with cd1:
                f_7d = st.checkbox("7 Dias", key="f_7d", on_change=_ao_marcar_7d)
            with cd2:
                f_14d = st.checkbox("14 Dias", key="f_14d", on_change=_ao_marcar_14d)
            with cd3:
                f_28d = st.checkbox("28 Dias", key="f_28d", on_change=_ao_marcar_28d)
            with cd4:
                f_datas = st.date_input(
                    "Personalizado",
                    value=(),
                    key="f_datas",
                    on_change=_ao_alterar_datas,
                    label_visibility="collapsed",
                )

        with col_ctr:
            st.markdown('<div style="height: 1.65rem"></div>', unsafe_allow_html=True)
            f_caixa = st.checkbox("Caixa Azul", key="f_caixa")

        with col_dir:
            f_cat = st.selectbox(
                "Categoria (Opcional)",
                options=CATEGORIAS,
                index=None,
                placeholder="Selecione a categoria",
                key="f_cat",
            )
            f_status = st.selectbox(
                "Status (Opcional)",
                options=["Armazenado", "Caixa Azul", "Museu", "A Doar", "Doados", "Descartados", "Devolvidos"],
                index=None,
                placeholder="Selecione o status",
                key="f_status",
            )

    dias_filtro = 7 if f_7d else (14 if f_14d else (28 if f_28d else None))

    df_filtrado = _aplicar_filtros(df.copy(), {
        "categoria": f_cat,
        "status": f_status,
        "caixa_azul": f_caixa,
        "texto": f_texto,
        "dias": dias_filtro,
        "datas": f_datas,
    })

    st.caption(f"{len(df_filtrado)} item(ns) encontrado(s)")

    if df_filtrado.empty:
        st.info("Nenhum item corresponde aos filtros.")
        return

    # ── Tabela ──
    df_editor = df_filtrado[
        ["codigo_item", "categoria", "descricao", "status_exibicao", "dias_no_status"]
    ].copy()
    df_editor.columns = ["Código", "Categoria", "Descrição", "Status", "Dias"]
    df_editor.insert(0, "✓", False)

    resultado = st.data_editor(
        df_editor,
        use_container_width=True,
        hide_index=True,
        column_config={
            "✓":         st.column_config.CheckboxColumn("✓", width="small"),
            "Código":    st.column_config.TextColumn("Código", width="small"),
            "Categoria": st.column_config.TextColumn("Categoria", width="medium"),
            "Descrição": st.column_config.TextColumn("Descrição", width="large"),
            "Status":    st.column_config.TextColumn("Status", width="medium"),
            "Dias":      st.column_config.NumberColumn("Dias", width="small"),
        },
        disabled=["Código", "Categoria", "Descrição", "Status", "Dias"],
        key="tabela_itens",
    )

    selecionados = df_filtrado["codigo_item"].iloc[
        resultado[resultado["✓"] == True].index.tolist()
    ].tolist()

    # ── Ações em lote e exportação ──
    st.markdown("---")
    col_lote, col_editar, col_exp_excel, col_exp_pdf = st.columns([2, 1, 1, 1])

    with col_lote:
        if perfil in PERFIS_EDICAO:
            if st.button(
                f"Movimentar ({len(selecionados)})" if selecionados else "Movimentar selecionados",
                disabled=len(selecionados) == 0,
                use_container_width=True,
                help="Selecione um ou mais itens",
                key="btn_lote",
            ):
                _dialog_lote(selecionados, perfil, usuario)

    with col_editar:
        if perfil in PERFIS_EDICAO:
            if st.button(
                "Editar",
                disabled=len(selecionados) != 1,
                use_container_width=True,
                help="Selecione um item",
                key="btn_editar",
            ):
                row_sel = df_filtrado[df_filtrado["codigo_item"] == selecionados[0]].iloc[0]
                _dialog_item(row_sel.to_dict(), perfil, usuario)

    df_sel = (
        df_filtrado[df_filtrado["codigo_item"].isin(selecionados)]
        if selecionados
        else pd.DataFrame(columns=COLUNAS_EXPORT)
    )

    with col_exp_excel:
        with st.popover("⬇ Exportar para Excel", use_container_width=True):
            st.download_button(
                "Exportar tudo",
                data=_exportar_excel(df_filtrado),
                file_name=f"inventario_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_exp_excel_tudo",
            )
            label_sel = f"Exportar seleção ({len(selecionados)})" if selecionados else "Exportar seleção"
            st.download_button(
                label_sel,
                data=_exportar_excel(df_sel),
                file_name=f"selecao_{date.today().isoformat()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="btn_exp_excel_sel",
            )

    with col_exp_pdf:
        with st.popover("⬇ Exportar para PDF", use_container_width=True):
            pdf_tudo = _exportar_pdf(df_filtrado)
            if pdf_tudo:
                st.download_button(
                    "Exportar tudo",
                    data=pdf_tudo,
                    file_name=f"inventario_{date.today().isoformat()}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_exp_pdf_tudo",
                )
            else:
                st.caption("Instale fpdf2 para habilitar PDF.")
            pdf_sel = _exportar_pdf(df_sel)
            if pdf_sel:
                label_sel_pdf = f"Exportar seleção ({len(selecionados)})" if selecionados else "Exportar seleção"
                st.download_button(
                    label_sel_pdf,
                    data=pdf_sel,
                    file_name=f"selecao_{date.today().isoformat()}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_exp_pdf_sel",
                )
            else:
                st.caption("Instale fpdf2 para habilitar PDF.")